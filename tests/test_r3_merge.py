"""QA del merge de R3 (sin browser):

1. `_merge_xlsx_files`: concat vertical de los chunks SIN filtrar — los accruals
   con fecha futura se conservan (Paul los quiere: son programados reales hasta
   el fondo del slicer). Sigue salteando la fila "Applied filters:" que inyecta
   Power BI y validando que todos los chunks tengan el mismo header.
2. `_chunk_date_range`: ventanas contiguas que cubren TODO el rango (esto es lo
   que garantiza "no se saltea ningun PA").
"""
import datetime as dt
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.scraper import _merge_xlsx_files, _chunk_date_range

HEADER = [
    "Client Name", "Client DDDID", "PA Number", "Start Date",
    "End Date", "Accrual Schedule Date", "Accrual Schedule Amount",
]


def _row(pa, accrual_dt, amount=100):
    return ["Cli", "1", pa, dt.datetime(2025, 6, 8),
            dt.datetime(2027, 1, 1), accrual_dt, amount]


def _make_xlsx(path, rows, header=HEADER, applied_filters_row=True):
    wb = Workbook()
    ws = wb.active
    ws.append(header)
    for r in rows:
        ws.append(r)
    if applied_filters_row:
        # PBI agrega esta fila al final de cada export; el merge debe saltearla.
        ws.append(["Applied filters: EndDate is on or after X"] + [None] * 6)
    wb.save(path)


def _read(path):
    wb = load_workbook(path, read_only=True)
    try:
        rows = list(wb.active.iter_rows(values_only=True))
    finally:
        wb.close()  # Windows: sin esto el handle bloquea el cleanup del tmpdir.
    return rows[0], rows[1:]


class MergeTests(unittest.TestCase):
    def setUp(self):
        self._tmp = tempfile.TemporaryDirectory()
        self.d = Path(self._tmp.name)

    def tearDown(self):
        self._tmp.cleanup()

    def test_keeps_future_accruals_and_all_pas(self):
        """No se filtra nada: pasados y futuros se conservan; PAs intactos. Se
        saltea solo la fila 'Applied filters:'."""
        p1 = self.d / "c1.xlsx"
        p2 = self.d / "c2.xlsx"
        _make_xlsx(p1, [
            _row("PA1", dt.datetime(2025, 8, 1)),     # pasado
            _row("PA1", dt.datetime(2026, 6, 2)),     # hoy-ish
            _row("PA1", dt.datetime(2026, 12, 1)),    # futuro -> se conserva
            _row("PA1", dt.datetime(2027, 7, 18)),    # futuro lejano -> se conserva
        ])
        _make_xlsx(p2, [
            _row("PA2", dt.datetime(2025, 9, 1)),
        ])
        out = self.d / "merged.xlsx"
        _merge_xlsx_files([p1, p2], out)

        header, data = _read(out)
        self.assertEqual(list(header), HEADER)
        self.assertEqual(len(data), 5)  # 4 de c1 + 1 de c2, ninguna borrada
        self.assertEqual({r[2] for r in data}, {"PA1", "PA2"})
        # Los accruals futuros estan presentes.
        accrual = {r[5].date() for r in data}
        self.assertIn(dt.date(2027, 7, 18), accrual)
        # La fila "Applied filters:" no debe aparecer.
        self.assertFalse(any(isinstance(r[0], str) and r[0].startswith("Applied filters")
                             for r in data))

    def test_header_mismatch_raises(self):
        p1 = self.d / "c1.xlsx"
        p2 = self.d / "c2.xlsx"
        _make_xlsx(p1, [_row("PA1", dt.datetime(2025, 8, 1))])
        _make_xlsx(p2, [_row("PA2", dt.datetime(2025, 9, 1))],
                   header=HEADER[:-1] + ["Otra Col"])
        out = self.d / "merged.xlsx"
        with self.assertRaises(ValueError):
            _merge_xlsx_files([p1, p2], out)

    def test_empty_paths_raises(self):
        with self.assertRaises(ValueError):
            _merge_xlsx_files([], self.d / "x.xlsx")


class ChunkRangeTests(unittest.TestCase):
    def test_contiguous_no_gaps_no_overlaps_covers_full_range(self):
        start, end, n = dt.date(2025, 6, 11), dt.date(2027, 7, 21), 8
        chunks = _chunk_date_range(start, end, n)
        self.assertEqual(len(chunks), n)
        self.assertEqual(chunks[0][0], start)          # arranca en start
        self.assertEqual(chunks[-1][1], end)           # termina en end (cubre todo)
        for i in range(1, n):
            prev_end = chunks[i - 1][1]
            cur_start = chunks[i][0]
            # contiguo y sin solape: el chunk siguiente arranca justo al dia despues.
            self.assertEqual(cur_start, prev_end + dt.timedelta(days=1),
                             f"gap/overlap entre chunk {i-1} y {i}: {chunks}")
            self.assertLessEqual(chunks[i][0], chunks[i][1])  # start <= end

    def test_invalid_n_raises(self):
        with self.assertRaises(ValueError):
            _chunk_date_range(dt.date(2025, 1, 1), dt.date(2025, 12, 31), 0)

    def test_range_too_small_raises(self):
        with self.assertRaises(ValueError):
            _chunk_date_range(dt.date(2025, 1, 1), dt.date(2025, 1, 3), 8)


if __name__ == "__main__":
    unittest.main()
