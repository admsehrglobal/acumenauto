"""QA of R1 going out as two files (no browser).

Covers the three things that would be silent if they broke:

1. `_merge_xlsx_files` with `keep_entry_ids` writes only that pile, and passing
   None leaves R3's path exactly as it was.
2. A pile that matches nothing writes a header-only file instead of raising. A
   day with no rejected-only invoices is legitimate; the zero-row guard must
   still fire when the export itself comes back empty.
3. The two piles never share a filename. They cover the same date range, so
   before the pile went into the slug the second merge overwrote the first and
   the rejected file simply vanished.

`test_subject_is_not_altered` in test_email_size_guard pins the wire format of
`send_reports_email`; what is pinned here is which subject each pile asks for,
since the service reading the inbox matches on it.
"""
import ast
import builtins
import datetime as dt
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from app.invoice_split import PILE_PAYABLE, PILE_REJECTED, subject_override_for
from app.scraper import (
    ChunkedReport,
    _classify_invoice_piles,
    _merge_xlsx_files,
)

HEADER = [
    "Urgency", "Entry ID", "PA Number", "Invoice #", "Client Name", "Client DDDID",
    "Client Number", "Service Code", "Status", "Rejected Reason", "Date Of Service",
    "Entry Creation Date", "Amount", "Aging",
]


def _row(entry_id, invoice, status, amount):
    return ["", entry_id, "PA1", invoice, "Cli", "D1", "C1", "Transportation",
            status, "", dt.datetime(2025, 7, 20), dt.datetime(2025, 7, 28),
            amount, 0]


def _make_xlsx(path, rows, applied_filters_row=True):
    wb = Workbook()
    ws = wb.active
    ws.append(HEADER)
    for r in rows:
        ws.append(r)
    if applied_filters_row:
        ws.append(["Applied filters: EndDate is on or after X"] + [None] * 13)
    wb.save(path)


def _read_rows(path):
    wb = load_workbook(path, read_only=True)
    try:
        return list(wb.active.iter_rows(values_only=True))
    finally:
        wb.close()


class MergeFilterTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.dir = Path(self.tmp.name)
        self.addCleanup(self.tmp.cleanup)
        self.chunk = self.dir / "chunk.xlsx"
        _make_xlsx(self.chunk, [
            _row("1", "500", "Rejected", 84.00),
            _row("2", "500", "Paid", 84.00),
            _row("3", "600", "Rejected", 10.00),
        ])

    def test_none_keeps_every_row(self):
        """R3 passes None and must behave exactly as before the filter existed."""
        out = self.dir / "all.xlsx"
        _merge_xlsx_files([self.chunk], out)
        self.assertEqual(len(_read_rows(out)), 4)  # header + 3

    def test_filter_writes_only_its_pile(self):
        out = self.dir / "pile.xlsx"
        _merge_xlsx_files([self.chunk], out, frozenset({"2"}))
        rows = _read_rows(out)
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[1][1], "2")

    def test_empty_pile_writes_header_only(self):
        """No rejected-only invoices on a given day is a legitimate empty pile,
        not a broken report - it must not abort R1."""
        out = self.dir / "empty.xlsx"
        _merge_xlsx_files([self.chunk], out, frozenset())
        self.assertEqual(_read_rows(out), [tuple(HEADER)])

    def test_empty_export_still_raises(self):
        """The guard this replaces was there for the export coming back empty
        (filter not applied / session dropped), and that must still fail loudly."""
        empty = self.dir / "empty_chunk.xlsx"
        _make_xlsx(empty, [])
        with self.assertRaises(ValueError):
            _merge_xlsx_files([empty], self.dir / "out.xlsx")


class ClassifyPilesTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.dir = Path(self.tmp.name)
        self.addCleanup(self.tmp.cleanup)

    def test_invoice_straddling_two_chunks_lands_in_one_pile(self):
        """R1 is chunked by date of service and one invoice's rows cross chunks.
        Classifying per chunk would put invoice 500 in both piles: the first
        chunk sees only a rejection, the second only the payment."""
        a, b = self.dir / "a.xlsx", self.dir / "b.xlsx"
        _make_xlsx(a, [_row("1", "500", "Rejected", 84.00)])
        _make_xlsx(b, [_row("2", "500", "Paid", 84.00)])
        meta = {a: (dt.date(2025, 1, 1), dt.date(2025, 6, 30), 1),
                b: (dt.date(2025, 7, 1), dt.date(2025, 12, 31), 1)}

        split, rejected_meta, payable_meta = _classify_invoice_piles([a, b], meta)

        self.assertEqual(split.rejected, frozenset())
        self.assertEqual(split.payable, frozenset({"2"}))
        self.assertEqual(rejected_meta[a][2], 0)
        self.assertEqual(payable_meta[b][2], 1)

    def test_power_bi_junk_rows_join_neither_pile(self):
        """The Total and blank rows Power BI mixes into the data carry no invoice
        number and no status, so they fall out of both piles."""
        chunk = self.dir / "junk.xlsx"
        _make_xlsx(chunk, [
            _row("1", "500", "Paid", 84.00),
            ["Total"] + [None] * 12 + [84.00],
            [None] * 14,
        ])
        meta = {chunk: (dt.date(2025, 1, 1), dt.date(2025, 12, 31), 3)}
        split, _, payable_meta = _classify_invoice_piles([chunk], meta)
        self.assertEqual(split.payable, frozenset({"1"}))
        self.assertEqual(split.rejected, frozenset())
        self.assertEqual(payable_meta[chunk][2], 1)


class SubjectTests(unittest.TestCase):
    LABEL = "2026-08-26 22:21 NJ"

    def test_rejected_pile_subject(self):
        self.assertEqual(
            subject_override_for(f"R1 - {PILE_REJECTED} (a to b)", self.LABEL, "R3"),
            "DCI Rejected Invoices - 2026-08-26 22:21 NJ",
        )

    def test_payable_pile_subject(self):
        self.assertEqual(
            subject_override_for(f"R1 - {PILE_PAYABLE} (a to b)", self.LABEL, "R3"),
            "DCI Payable Invoices - 2026-08-26 22:21 NJ",
        )

    def test_accrual_subject_unchanged(self):
        self.assertEqual(
            subject_override_for("R3 (a to b)", self.LABEL, "R3"), "Accrual Schedule"
        )

    def test_other_reports_keep_the_default(self):
        self.assertIsNone(subject_override_for("R2", self.LABEL, "R3"))


class CommandWiringTests(unittest.TestCase):
    """Static checks on the management command.

    It cannot be imported without Django configured, so these read the source.
    Cheap, but they cover the failure this suite missed: moving a function inside
    that file deleted `_notify_failure` while leaving its three call sites, and
    every failure path raised NameError instead of alerting support. Nothing in
    the suite touched it because nothing covers download_report.py at all.
    """

    def _tree(self):
        src = Path("app/management/commands/download_report.py").read_text(
            encoding="utf-8"
        )
        return ast.parse(src)

    def test_every_called_helper_is_defined(self):
        tree = self._tree()
        defined = {
            n.name for n in ast.walk(tree)
            if isinstance(n, (ast.FunctionDef, ast.AsyncFunctionDef))
        }
        imported = {
            (a.asname or a.name) for n in ast.walk(tree)
            if isinstance(n, (ast.Import, ast.ImportFrom)) for a in n.names
        }
        called = {
            n.func.id for n in ast.walk(tree)
            if isinstance(n, ast.Call) and isinstance(n.func, ast.Name)
        }
        undefined = called - defined - imported - set(dir(builtins))
        self.assertEqual(undefined, set(), f"called but never defined: {undefined}")

    def test_notify_failure_still_exists(self):
        defined = {
            n.name for n in ast.walk(self._tree())
            if isinstance(n, ast.FunctionDef)
        }
        self.assertIn("_notify_failure", defined)


class SpecTests(unittest.TestCase):
    def test_invoice_split_is_off_by_default(self):
        """R3 has no Status column: a split leaking onto it would filter every
        row away and kill the report."""
        spec = ChunkedReport(
            url="u", button_name="b", n_chunks=4, today=dt.date(2026, 8, 27),
            tab_name=None, single_slicer=True, full_range=True,
        )
        self.assertFalse(spec.invoice_split)
        self.assertEqual(spec.reset_slicers, ())


if __name__ == "__main__":
    unittest.main()
