"""QA of `_split_for_email`: how a merged report becomes one or several emails.

One report is one attachment while it fits. Above the Brevo ceiling the chunks
are regrouped into consecutive date ranges, one file per email — Paul confirmed
on 2026-07-27 that the service reading the inbox takes several separate emails.
What must hold in every case: no row is lost, the ranges stay contiguous, and
each file is labelled with the range it actually carries.

The real ceiling is ~15MB, so the constants are patched to the size of the
fixture instead of writing 15MB of xlsx.
"""
import datetime as dt
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook, load_workbook

from app import scraper
from app.scraper import _merge_xlsx_files, _split_for_email

HEADER = ["Client Name", "PA Number", "Date of Service", "Amount"]
D0 = dt.date(2025, 1, 1)


def _make_xlsx(path: Path, tags: list[str]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(HEADER)
    for tag in tags:
        ws.append(["Cliente con nombre largo repetido", tag,
                   dt.datetime(2025, 6, 8), 123.45])
    # PBI appends this row to every export; the merge skips it.
    ws.append(["Applied filters: DateOfService is on or after X", None, None, None])
    wb.save(path)


def _pas(path: Path) -> list[str]:
    wb = load_workbook(path, read_only=True)
    try:
        rows = list(wb.active.iter_rows(values_only=True))
    finally:
        wb.close()  # Windows: sin esto el handle bloquea el cleanup del tmpdir.
    return [r[1] for r in rows[1:]]


class SplitForEmailTests(unittest.TestCase):
    def setUp(self):
        self._tmp = tempfile.TemporaryDirectory()
        self.d = Path(self._tmp.name)
        # 4 chunks of 25 rows, consecutive 10-day ranges starting at D0.
        self.parts = []
        self.part_meta = {}
        for i in range(4):
            p = self.d / f"r1_part_{i:03d}.xlsx"
            _make_xlsx(p, [f"P{i}-{j}" for j in range(25)])
            start = D0 + dt.timedelta(days=10 * i)
            self.part_meta[p] = (start, start + dt.timedelta(days=9), 25)
            self.parts.append(p)
        self.report_range = (D0, D0 + dt.timedelta(days=39))
        self.merged = self.d / "r1_2025-01-01_to_2025-02-09_ts.xlsx"
        _merge_xlsx_files(self.parts, self.merged)

    def tearDown(self):
        self._tmp.cleanup()

    def _split(self, max_bytes, target_bytes):
        with patch.object(scraper, "_ATTACHMENT_MAX_BYTES", max_bytes), \
                patch.object(scraper, "_ATTACHMENT_TARGET_BYTES", target_bytes):
            return _split_for_email(
                self.merged, self.parts, self.part_meta, self.report_range,
                self.d, "r1", "ts",
            )

    def test_single_file_while_it_fits(self):
        """Un reporte que entra en un mail sale como hoy: un solo adjunto, con
        el rango completo (o sea, el mismo subject de siempre)."""
        size = self.merged.stat().st_size
        outputs = self._split(size * 2, size)
        self.assertEqual(outputs, [(self.merged, *self.report_range)])
        self.assertTrue(self.merged.exists())

    def test_splits_into_contiguous_files_without_losing_rows(self):
        size = self.merged.stat().st_size
        expected = _pas(self.merged)
        # target = un cuarto del merge -> ~25 filas por archivo -> 4 archivos.
        outputs = self._split(size // 2, size // 4)

        self.assertGreater(len(outputs), 1)
        # Ni una fila de menos, ni duplicadas, y en el mismo orden.
        got = [pa for path, _, _ in outputs for pa in _pas(path)]
        self.assertEqual(got, expected)
        # Rangos contiguos que cubren exactamente el rango del reporte.
        self.assertEqual(outputs[0][1], self.report_range[0])
        self.assertEqual(outputs[-1][2], self.report_range[1])
        for prev, cur in zip(outputs, outputs[1:]):
            self.assertEqual(cur[1], prev[2] + dt.timedelta(days=1))
        # Cada archivo lleva su rango en el nombre (va al subject del mail).
        for path, first, last in outputs:
            self.assertIn(f"{first.isoformat()}_to_{last.isoformat()}", path.name)
        # El merge grande se borra: no se manda dos veces la misma data.
        self.assertFalse(self.merged.exists())

    def test_a_single_oversized_chunk_is_left_alone(self):
        """No se puede partir mas fino que un chunk: se devuelve el merge tal
        cual y el guard de send_reports_email falla ese mail con el motivo."""
        one = self.parts[:1]
        merged = self.d / "solo.xlsx"
        _merge_xlsx_files(one, merged)
        with patch.object(scraper, "_ATTACHMENT_MAX_BYTES", 1), \
                patch.object(scraper, "_ATTACHMENT_TARGET_BYTES", 1):
            outputs = _split_for_email(
                merged, one, self.part_meta, self.report_range, self.d, "r1", "ts",
            )
        self.assertEqual(outputs, [(merged, *self.report_range)])
        self.assertTrue(merged.exists())

    def test_empty_chunks_never_form_a_group_of_their_own(self):
        """Los sub-rangos vacios que genera el chunking adaptativo viajan con el
        grupo anterior: un grupo sin filas haria fallar el merge."""
        empty = self.d / "r1_part_009.xlsx"
        _make_xlsx(empty, [])
        last_end = self.part_meta[self.parts[-1]][1]
        self.part_meta[empty] = (
            last_end + dt.timedelta(days=1), last_end + dt.timedelta(days=10), 0,
        )
        parts = self.parts + [empty]
        report_range = (D0, last_end + dt.timedelta(days=10))
        merged = self.d / "con_vacio.xlsx"
        _merge_xlsx_files(parts, merged)
        size = merged.stat().st_size

        with patch.object(scraper, "_ATTACHMENT_MAX_BYTES", size // 2), \
                patch.object(scraper, "_ATTACHMENT_TARGET_BYTES", size // 5):
            outputs = _split_for_email(
                merged, parts, self.part_meta, report_range, self.d, "r1", "ts",
            )

        self.assertGreater(len(outputs), 1)
        self.assertEqual(outputs[-1][2], report_range[1])


if __name__ == "__main__":
    unittest.main()
