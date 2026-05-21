"""DCI portal scraper.

Flow: login en acumen.dcisoftware.com -> click "Advanced Insights" abre popup
que hace handshake OAuth con acumen-xcore-auth (hay que esperarlo para que
setee cookies) -> por cada reporte: navega al report group -> click boton
del reporte -> exporta el iframe de Power BI a Excel.

El reporte 3 (Vendor Authorization Accrual Balances) excede 150k filas/export
asi que se baja en N chunks: click una sola vez al boton, despues loop
seteando el date filter (textbox Angular Material) y exportando cada rango.
Los N chunks se mergean en un unico xlsx antes de devolver, para que el
cliente reciba un solo adjunto en lugar de N emails separados.

Selectores relevados con `playwright codegen` (reporte 1: 2026-04-24,
reporte 2: 2026-04-27, reporte 3: 2026-05-10).
"""
from __future__ import annotations

import asyncio
import datetime as dt
import logging
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from playwright.async_api import (
    BrowserContext,
    Page,
    TimeoutError as PlaywrightTimeoutError,
    async_playwright,
)

logger = logging.getLogger(__name__)

PORTAL_URL = "https://acumen.dcisoftware.com/"


async def download_reports(
    username: str,
    password: str,
    reports: list[tuple[str, str]],
    output_dir: Path,
    timestamp_label: str,
    chunked_report: tuple[str, str, int, dt.date] | None = None,
) -> list[tuple[Path, str]]:
    """Login una vez, descarga cada reporte reusando el popup.

    `reports` es la lista de reportes simples como (report_url, button_name).
    `chunked_report` es el reporte que excede el limite de filas por export y
    se descarga en N chunks: (url, button_name, n_chunks, today).
    El rango de fechas se determina asi: start_date = lo mas anterior que
    permita el PBI (MIN del slicer Accrual Schedule Date, leido del aria-label);
    end_date = `today` (hoy en NJ tz, lo pasa el caller), clamped al MAX del
    slicer si today va mas alla.
    `timestamp_label` se appendea al nombre del archivo para que cada run quede
    identificable (ej: '2026-04-27_14h30').

    Devuelve lista de (path, display_name) por archivo descargado — un tuple
    por reporte simple, N tuples por reporte chunked. El display_name va al
    subject del email.
    """
    output_dir.mkdir(parents=True, exist_ok=True)

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(
            accept_downloads=True,
            locale="en-US",
        )
        # Las paginas disparan una cadena federada de OAuth (xcore -> xcore-auth
        # -> portal principal -> vuelta) que puede tardar > 30s default.
        context.set_default_timeout(60000)

        try:
            page = await context.new_page()
            await _login(page, username, password)
            report_page = await _open_reports_popup(page, username, password)

            results: list[tuple[Path, str]] = []
            for report_url, button_name in reports:
                await report_page.goto(report_url)
                logger.warning("[REPORT] URL post-goto: %s", report_page.url)
                path = await _export_excel(
                    report_page, button_name, output_dir, timestamp_label
                )
                results.append((path, button_name))

            if chunked_report is not None:
                url, button_name, n_chunks, today = chunked_report
                await report_page.goto(url)
                logger.warning("[REPORT chunked] URL post-goto: %s", report_page.url)
                results.extend(
                    await _export_chunked_report(
                        report_page,
                        button_name,
                        n_chunks,
                        output_dir,
                        timestamp_label,
                        today,
                    )
                )
            return results
        except Exception:
            await _dump_debug(context, output_dir)
            raise
        finally:
            await context.close()
            await browser.close()


async def _login(page: Page, username: str, password: str) -> None:
    await page.goto(PORTAL_URL)

    # Modal opcional que aparece a veces pre-login.
    try:
        await page.get_by_text("× Close Sign In Fake Username").click(timeout=2000)
    except PlaywrightTimeoutError:
        pass

    await page.get_by_role("textbox", name="Username*").fill(username)
    await page.get_by_role("textbox", name="Password*").fill(password)
    await page.get_by_role("button", name="Sign In").click()

    await page.get_by_role("link", name="Advanced Insights").wait_for()


async def _open_reports_popup(page: Page, username: str, password: str) -> Page:
    async with page.expect_popup() as popup_info:
        await page.get_by_role("link", name="Advanced Insights").click()
    popup = await popup_info.value
    await popup.wait_for_load_state("load")
    logger.warning("[POPUP] URL post-load: %s", popup.url)
    # A veces el popup autentica via SSO silencioso; a veces cae en el IdP de
    # xcore (acumen-xcore-auth) pidiendo credenciales de nuevo. En ese caso
    # llenamos el form — es una pantalla distinta al portal principal.
    if "/Account/Login" in popup.url:
        logger.warning("[POPUP] Cayó en Account/Login, haciendo XCore login")
        await _xcore_login(popup, username, password)
        logger.warning("[POPUP] URL post-xcore-login: %s", popup.url)
    else:
        logger.warning("[POPUP] SSO silencioso OK")
    return popup


async def _xcore_login(page: Page, username: str, password: str) -> None:
    await page.get_by_placeholder("User name or email address").fill(username)
    await page.get_by_placeholder("Password").fill(password)
    await page.get_by_role("button", name="Login").click()
    await page.wait_for_url(lambda url: "/Account/Login" not in url)


async def _export_excel(
    page: Page, report_button_name: str, output_dir: Path, timestamp_label: str
) -> Path:
    await page.get_by_role("button", name=report_button_name).click()

    # DIAG: capturar estado post-click para saber si la pagina hung, o si el
    # iframe esta presente con otro title. Borrar despues de diagnosticar.
    try:
        await page.wait_for_load_state("networkidle", timeout=10000)
        networkidle = "ok"
    except PlaywrightTimeoutError:
        networkidle = "timeout-10s"
    all_iframe_titles = await page.locator("iframe").evaluate_all(
        "els => els.map(e => e.getAttribute('title'))"
    )
    logger.warning(
        "[DIAG export_excel] post-click button=%r url=%s networkidle=%s iframes=%r",
        report_button_name, page.url, networkidle, all_iframe_titles,
    )

    iframe_element = page.locator('iframe[title="Embedded report"]')
    # 180s en lugar del default de 60s: Power BI a veces tarda > 60s en insertar
    # el iframe en el DOM post-click (confirmado 2026-05-21 con DIAG: iframe=[]
    # a los 23s post-click, recien aparecio entre 60-90s). Default mataba el
    # cron daily intermitentemente.
    try:
        await iframe_element.wait_for(timeout=180000)
    except PlaywrightTimeoutError:
        final_titles = await page.locator("iframe").evaluate_all(
            "els => els.map(e => e.getAttribute('title'))"
        )
        logger.error(
            "[DIAG export_excel] wait_for FAILED. final url=%s iframes=%r",
            page.url, final_titles,
        )
        raise
    # Hover sobre el iframe fuerza que Power BI muestre el menú "..." del visual.
    # Sin esto, en headless el boton visual-more-options-btn puede no renderizarse.
    await iframe_element.hover()
    iframe = iframe_element.content_frame

    more_btn = iframe.get_by_test_id("visual-more-options-btn")
    await more_btn.wait_for(state="attached")
    # force=True evita que tooltips de Power BI intercepten los clicks en headless.
    await more_btn.click(force=True)
    await iframe.get_by_test_id("pbimenu-item.Export data").click(force=True)
    await iframe.get_by_text("Data with current layout").click(force=True)

    async with page.expect_download() as download_info:
        await iframe.get_by_test_id("export-btn").click(force=True)
    download = await download_info.value

    # El portal siempre sugiere "data.xlsx" — derivamos del button_name + timestamp
    # para no pisar archivos y que cada run quede identificable en el inbox.
    slug = "_".join(report_button_name.lower().split())
    target = output_dir / f"{slug}_{timestamp_label}.xlsx"
    await download.save_as(target)
    return target


async def _export_chunked_report(
    page: Page,
    button_name: str,
    n_chunks: int,
    output_dir: Path,
    timestamp_label: str,
    today: dt.date,
) -> list[tuple[Path, str]]:
    """Click el boton del reporte una vez y exporta N veces cambiando el rango
    (sin recargar la pagina entre chunks).

    El tab tiene 2 date range slicers: izq filtra por PA Start Date, der
    filtra por Accrual Schedule Date. Chunkeamos por el der → cada accrual
    cae en un solo chunk. Identificamos el slicer der parseando el aria-label
    `Available input range MIN to MAX` (pickeamos el par con MAX mas lejano
    en el futuro — el Accrual Schedule Date extent siempre va mas alla del
    PA Start Date extent porque incluye accruals futuros programados).
    El selector `.last` no es estable: el DOM order de los slicers varia
    entre runs segun timing de render de PBI (confirmado 2026-05-21 — en
    runs sucesivos `.last` agarro tanto slicer A como slicer B).

    start_date = MIN del slicer der (lo mas anterior que permite PBI).
    end_date = `today` (hoy NJ), clamped al MAX del slicer si fuera necesario.

    Los N chunks se mergean en un unico xlsx al final. Si cualquier chunk
    falla, la excepcion propaga sin mergear (abort-on-fail: el cliente no
    recibe un archivo parcial).
    """
    await page.get_by_role("button", name=button_name).click()

    iframe_element = page.locator('iframe[title="Embedded report"]')
    # Mismo motivo que en _export_excel: 180s para tolerar inserciones lentas
    # del iframe por parte de Power BI.
    await iframe_element.wait_for(timeout=180000)
    await iframe_element.hover()
    iframe = iframe_element.content_frame

    # El reporte abre por default en el tab 'Estimated Accrual Balances' (que
    # solo tiene totales). Switch al tab con detalle PA + schedule semanal.
    await iframe.get_by_role("tab", name="PA Details and Schedule by").click()

    # Slicer A (PA Start Date) suele cargar primero; B (Accrual Schedule Date)
    # carga unos segundos despues. Si leemos antes de que B este, identificamos
    # mal y caemos en A. Esperar hasta tener los 4 inputs date (2 slicers x 2
    # textboxes cada uno) garantiza que ambos esten antes de identificar.
    date_inputs = iframe.locator(
        "input[aria-label*='Available input range']"
    )
    deadline = asyncio.get_event_loop().time() + 30
    while True:
        count = await date_inputs.count()
        if count >= 4:
            break
        if asyncio.get_event_loop().time() > deadline:
            raise TimeoutError(
                f"Only {count} date inputs after 30s — slicer B nunca cargo"
            )
        await asyncio.sleep(0.5)

    start_idx, end_idx, slicer_min, slicer_max, date_fmt = (
        await _identify_accrual_slicer(date_inputs)
    )
    start_date = slicer_min
    end_date = min(today, slicer_max)
    logger.warning(
        "[REPORT chunked] Slicer identified: start_idx=%d end_idx=%d "
        "range=%s to %s (fmt %s)",
        start_idx, end_idx, slicer_min, slicer_max, date_fmt,
    )
    logger.warning(
        "[REPORT chunked] Effective range: %s to %s (today=%s, clamped to slicer max=%s)",
        start_date, end_date, today, slicer_max,
    )

    start_input = date_inputs.nth(start_idx)
    end_input = date_inputs.nth(end_idx)

    chunks = _chunk_date_range(start_date, end_date, n_chunks)
    slug = "_".join(button_name.lower().split())
    part_paths: list[Path] = []

    for i, (chunk_start, chunk_end) in enumerate(chunks, start=1):
        chunk_label = (
            f"Part {i}/{n_chunks} ({chunk_start.isoformat()} to "
            f"{chunk_end.isoformat()})"
        )
        logger.warning("[REPORT chunked] Exporting %s", chunk_label)

        await _set_date_filter(
            start_input, end_input, chunk_start, chunk_end, date_fmt
        )

        # El tab nuevo tiene multiples visuals — scope al table visual via
        # aria-label ("Row" lo distingue de los charts).
        table_visual = iframe.get_by_role("group").filter(
            has_text="Scroll left Scroll right Row"
        )
        more_btn = table_visual.get_by_test_id("visual-more-options-btn")
        export_item = iframe.get_by_test_id("pbimenu-item.Export data")

        # En iter >= 2 a veces el click sobre "..." no abre el menu (el visual
        # esta busy con el re-render del filtro nuevo). Retry con Escape + hover
        # entre intentos para limpiar el estado.
        for attempt in range(3):
            await page.keyboard.press("Escape")
            await asyncio.sleep(0.5)
            await table_visual.hover()
            await more_btn.wait_for(state="visible")
            await more_btn.click(force=True)
            try:
                await export_item.wait_for(state="visible", timeout=5000)
                break
            except PlaywrightTimeoutError:
                logger.warning(
                    "[REPORT chunked] Menu didn't open on attempt %d/3, retrying",
                    attempt + 1,
                )
                if attempt == 2:
                    raise

        await export_item.click(force=True)
        await iframe.get_by_text("Data with current layout").click(force=True)

        async with page.expect_download() as download_info:
            await iframe.get_by_test_id("export-btn").click(force=True)
        download = await download_info.value

        part_path = output_dir / (
            f"{slug}_part_{i}_of_{n_chunks}"
            f"_{chunk_start.isoformat()}_to_{chunk_end.isoformat()}"
            f"_{timestamp_label}.xlsx"
        )
        await download.save_as(part_path)
        rows = _validate_chunk_xlsx(part_path)
        logger.warning(
            "[REPORT chunked] %s OK (%d data rows)", chunk_label, rows
        )
        part_paths.append(part_path)

    merged_path = output_dir / (
        f"{slug}_{start_date.isoformat()}_to_{end_date.isoformat()}"
        f"_{timestamp_label}.xlsx"
    )
    _merge_xlsx_files(part_paths, merged_path)
    for p in part_paths:
        p.unlink(missing_ok=True)

    display_name = (
        f"{button_name} ({start_date.isoformat()} to {end_date.isoformat()})"
    )
    return [(merged_path, display_name)]


def _validate_chunk_xlsx(path: Path) -> int:
    """Confirma que el chunk recien descargado es un xlsx valido con datos.

    Atrapa los modos de falla silenciosos mas comunes: descarga truncada
    (openpyxl no puede abrir), filtro no aplicado o sesion caida (export
    vacio o sin data rows). Si algo falla, raise -> el reporte aborta antes
    de seguir con los proximos chunks o el merge.

    Devuelve la cantidad de data rows (excluyendo header) para logging.
    """
    wb = load_workbook(path, read_only=True)
    try:
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            raise ValueError(f"{path.name}: archivo vacio (sin header)")
        if not any(cell is not None for cell in header):
            raise ValueError(f"{path.name}: header row vacio")
        data_rows = sum(1 for _ in rows_iter)
        if data_rows == 0:
            raise ValueError(f"{path.name}: sin data rows (solo header)")
        return data_rows
    finally:
        wb.close()


def _merge_xlsx_files(paths: list[Path], output_path: Path) -> Path:
    """Concat vertical de N xlsx con single-row header.

    El tab 'PA Details and Schedule by Client' devuelve tabla plana (header
    de 1 fila, columnas fijas). Como chunkeamos por Accrual Schedule Date,
    cada accrual aparece en exactamente un chunk (sin overlap). Concat vertical
    directo; las rows se preservan tal cual.

    Validamos que todos los chunks tengan el mismo header (Power BI siempre
    devuelve las mismas columnas para el mismo visual, independiente del
    filtro de fechas).
    """
    if not paths:
        raise ValueError("No paths to merge")

    # write_only=True: stream-escribe a disco en lugar de mantener todas las
    # celdas en RAM. Para 218k filas la diferencia es ~600MB vs ~11MB peak,
    # critico en el worker de Fly (2GB) donde Chromium aun corre en paralelo
    # y la presion de memoria hace que out_wb.save() default tarde 4+ min.
    canonical_header: tuple | None = None
    out_wb = Workbook(write_only=True)
    out_ws = out_wb.create_sheet()
    total_rows = 0

    for path in paths:
        wb = load_workbook(path, read_only=True)
        try:
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header = next(rows_iter)
            except StopIteration:
                raise ValueError(f"{path.name}: archivo vacio")
            if canonical_header is None:
                canonical_header = header
                out_ws.append(list(header))
            elif header != canonical_header:
                raise ValueError(
                    f"{path.name}: header no matchea con el primer chunk "
                    f"({header!r} vs {canonical_header!r})"
                )
            for row in rows_iter:
                # Power BI inyecta una fila al final de cada export con el
                # filtro aplicado (ej: "Applied filters: EndDate is on or
                # after X and is before Y"). La salteamos para que el archivo
                # final tenga solo data.
                first = row[0] if row else None
                if isinstance(first, str) and first.startswith("Applied filters:"):
                    continue
                out_ws.append(list(row))
                total_rows += 1
        finally:
            wb.close()

    logger.warning(
        "[REPORT chunked] Merged %d files -> %s (%d data rows, %d cols)",
        len(paths), output_path.name, total_rows,
        len(canonical_header) if canonical_header else 0,
    )
    out_wb.save(output_path)
    return output_path


def _chunk_date_range(
    start: dt.date, end: dt.date, n: int
) -> list[tuple[dt.date, dt.date]]:
    """Parte [start, end] en N rangos contiguos sin overlaps ni gaps."""
    if n < 1:
        raise ValueError(f"n_chunks debe ser >= 1, got {n}")
    total_days = (end - start).days
    if total_days < n:
        raise ValueError(
            f"Rango muy chico para {n} chunks: {start} to {end} ({total_days} días)"
        )
    size = total_days / n
    out: list[tuple[dt.date, dt.date]] = []
    for i in range(n):
        cs = start + dt.timedelta(days=int(i * size))
        ce = end if i == n - 1 else start + dt.timedelta(days=int((i + 1) * size) - 1)
        out.append((cs, ce))
    return out


def _parse_filter_date(s: str) -> tuple[dt.date, str]:
    """Parsea el string del date filter de Power BI y devuelve (fecha, formato)
    para que al escribir de vuelta usemos el mismo formato que muestra el
    portal (Acumen podria estar en US o EU dependiendo del tenant)."""
    s = s.strip()
    for fmt in ("%m/%d/%Y", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(s, fmt).date(), fmt
        except ValueError:
            pass
    raise ValueError(f"Formato de fecha no reconocido: {s!r}")


async def _set_date_filter(
    start_input,
    end_input,
    start_date: dt.date,
    end_date: dt.date,
    date_fmt: str,
) -> None:
    """Escribe directo en los textbox del filtro Angular Material.

    Orden critico: END PRIMERO, despues START. Power BI valida start <= end al
    commitear cada campo y rechaza silenciosamente si falla. Como iteramos los
    chunks forward (start del chunk N+1 > end del chunk N), si seteamos start
    primero queda > end actual y se rechaza. Seteando end primero expandimos
    la ventana hacia adelante, despues podemos mover start sin violar la regla.

    El slicer commitea por debounce automatico (no necesita Tab/Enter) pero
    necesita un sleep para que el commit alcance a procesarse antes del
    proximo cambio o de exportar.

    `start_input` y `end_input` son locators ya posicionados (via .nth dentro
    de los date inputs del iframe). El aria-label se actualiza al commitear
    un filtro (la parte "range MIN to MAX" reproduce la seleccion actual), por
    eso el caller no debe pasar selectores por aria-label exacto."""
    await end_input.click()
    await end_input.fill(end_date.strftime(date_fmt))
    await asyncio.sleep(2)

    await start_input.click()
    await start_input.fill(start_date.strftime(date_fmt))

    # Power BI no expone una señal explicita de "filtro aplicado, visual
    # listo". Esperar un poco evita exportar mientras el visual aun esta
    # re-rendereando con los datos viejos.
    await asyncio.sleep(3)


_LABEL_RANGE_RE = re.compile(r"Available input range (\S+) to (\S+)$")


async def _identify_accrual_slicer(
    date_inputs,
) -> tuple[int, int, dt.date, dt.date, str]:
    """Identifica el slicer 'Accrual Schedule Date' entre los dos date range
    slicers del tab 'PA Details and Schedule by Client'.

    Cada slicer expone su rango disponible en el aria-label de sus 2 textbox
    ("Start date. Available input range MIN to MAX" / "End date..."). El
    slicer Accrual Schedule Date es el que tiene MAX mas lejano en el futuro
    (incluye accruals programados a futuro, mientras que el slicer PA Start
    Date no va mas alla del PA mas reciente).

    `date_inputs` es el locator `input[aria-label*="Available input range"]`
    del iframe — todos los inputs date de los slicers. Devolvemos los INDICES
    de los inputs de slicer B dentro de ese locator (no los aria-labels), para
    que el caller pueda usar `date_inputs.nth(idx)` como selector estable —
    el aria-label se mueve cuando aplicamos un filtro, las posiciones no."""
    inputs = await date_inputs.evaluate_all(
        "els => els.map((e, i) => ({i: i, label: e.getAttribute('aria-label')}))"
    )
    grouped: dict[tuple[str, str], dict[str, int]] = {}
    for item in inputs:
        label = item["label"] or ""
        m = _LABEL_RANGE_RE.search(label)
        if not m:
            continue
        if label.startswith("Start date."):
            kind = "start"
        elif label.startswith("End date."):
            kind = "end"
        else:
            continue
        grouped.setdefault((m.group(1), m.group(2)), {})[kind] = item["i"]

    pairs = []
    date_fmt = None
    for (min_s, max_s), idxs in grouped.items():
        if "start" not in idxs or "end" not in idxs:
            continue
        min_d, fmt = _parse_filter_date(min_s)
        max_d, _ = _parse_filter_date(max_s)
        date_fmt = fmt
        pairs.append((max_d, min_d, idxs["start"], idxs["end"]))

    if not pairs:
        raise ValueError(
            f"No date-range slicer pairs found. Inputs were: {inputs!r}"
        )

    # Mas lejano en el futuro = Accrual Schedule Date (slicer der).
    pairs.sort(key=lambda x: x[0], reverse=True)
    max_d, min_d, start_idx, end_idx = pairs[0]
    return start_idx, end_idx, min_d, max_d, date_fmt


async def _dump_debug(context: BrowserContext, output_dir: Path) -> None:
    """En error, dump screenshot + HTML + URL de cada pagina del context."""
    for i, p in enumerate(context.pages):
        try:
            await p.screenshot(
                path=str(output_dir / f"error_{i}.png"), full_page=True
            )
            (output_dir / f"error_{i}.url").write_text(
                f"{p.url}\n{await p.title()}\n", encoding="utf-8"
            )
            html = await p.content()
            (output_dir / f"error_{i}.html").write_text(html, encoding="utf-8")
            logger.error("Debug dump page %d: %s", i, p.url)
        except Exception as exc:
            logger.error("No pude capturar page %d: %s", i, exc)
